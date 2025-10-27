import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import os
import tempfile
from utils import format_currency, apply_excel_styles

class ReportGenerator:
    def __init__(self, input_file):
        # Try different encodings and handle CSV parsing issues
        encodings = ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']
        last_error = None
        
        for encoding in encodings:
            try:
                # Read CSV with explicit parameters for better parsing
                self.df = pd.read_csv(
                    input_file,
                    encoding=encoding,
                    skipinitialspace=True,  # Skip spaces after delimiter
                    on_bad_lines='warn',    # Warn about problematic lines
                    delimiter=',',          # Explicitly set delimiter
                )
                
                # Verify expected columns are present
                expected_columns = ['Date', 'Category', 'Product', 'Sales', 'Quantity']
                missing_columns = [col for col in expected_columns if col not in self.df.columns]
                
                if missing_columns:
                    raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
                
                # Convert date and numeric columns
                self.df['Date'] = pd.to_datetime(self.df['Date'])
                self.df['Sales'] = pd.to_numeric(self.df['Sales'], errors='coerce')
                self.df['Quantity'] = pd.to_numeric(self.df['Quantity'], errors='coerce')
                
                # Remove any rows with NaN values
                self.df = self.df.dropna()
                break
                
            except UnicodeDecodeError:
                continue
            except Exception as e:
                last_error = str(e)
                continue
        
        if self.df is None:
            raise Exception(f"Failed to read CSV file with any encoding. Last error: {last_error}")
            
        if len(self.df) == 0:
            raise Exception("No valid data found in the CSV file after processing")
        
    def create_pivot_tables(self):
        # Sales by Category
        category_sales = pd.pivot_table(
            self.df,
            values=['Sales', 'Quantity'],
            index=['Category'],
            aggfunc='sum'
        )
        
        # Sales by Date
        daily_sales = pd.pivot_table(
            self.df,
            values='Sales',
            index=['Date'],
            columns=['Category'],
            aggfunc='sum',
            fill_value=0
        )
        
        return category_sales, daily_sales
    
    def create_charts(self):
        import tempfile
        import os
        
        chart_paths = []
        
        try:
            # Category Sales Pie Chart
            plt.figure(figsize=(8, 6))
            self.df.groupby('Category')['Sales'].sum().plot(
                kind='pie',
                autopct='%1.1f%%'
            )
            plt.title('Sales by Category')
            
            # Save to temp file
            temp_file1 = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            plt.savefig(temp_file1.name, format='png', dpi=300, bbox_inches='tight')
            chart_paths.append(temp_file1.name)
            plt.close()
            
            # Daily Sales Line Chart
            plt.figure(figsize=(10, 6))
            daily_sales = self.df.groupby('Date')['Sales'].sum()
            daily_sales.plot(kind='line', marker='o')
            plt.title('Daily Sales Trend')
            plt.grid(True)
            
            # Save to temp file
            temp_file2 = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            plt.savefig(temp_file2.name, format='png', dpi=300, bbox_inches='tight')
            chart_paths.append(temp_file2.name)
            plt.close()
            
            return chart_paths
            
        except Exception as e:
            # Clean up any temporary files
            for path in chart_paths:
                try:
                    os.unlink(path)
                except:
                    pass
            raise Exception(f"Error creating charts: {str(e)}")
    
    def calculate_summary_stats(self):
        stats = {
            'Total Sales': self.df['Sales'].sum(),
            'Average Sale': self.df['Sales'].mean(),
            'Total Items Sold': self.df['Quantity'].sum(),
            'Unique Products': self.df['Product'].nunique(),
            'Date Range': f"{self.df['Date'].min().strftime('%Y-%m-%d')} to {self.df['Date'].max().strftime('%Y-%m-%d')}"
        }
        return stats
    
    def generate_report(self, output_file, include_pivot=True, 
                       include_charts=True, include_stats=True):
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Raw Data Sheet
        ws_data = wb.create_sheet("Raw Data")
        for r in dataframe_to_rows(self.df, index=False, header=True):
            ws_data.append(r)
        
        apply_excel_styles(ws_data)
        
        if include_pivot:
            category_sales, daily_sales = self.create_pivot_tables()
            
            # Category Sales Pivot
            ws_category = wb.create_sheet("Category Sales")
            for r in dataframe_to_rows(category_sales, index=True, header=True):
                ws_category.append(r)
            
            apply_excel_styles(ws_category)
            
            # Daily Sales Pivot
            ws_daily = wb.create_sheet("Daily Sales")
            for r in dataframe_to_rows(daily_sales, index=True, header=True):
                ws_daily.append(r)
            
            apply_excel_styles(ws_daily)
        
        if include_charts:
            try:
                chart_paths = self.create_charts()
                ws_charts = wb.create_sheet("Charts")
                
                # Add each chart to the worksheet
                for idx, chart_path in enumerate(chart_paths):
                    try:
                        img = Image(chart_path)
                        cell_pos = f'B{idx * 20 + 2}'
                        ws_charts.add_image(img, cell_pos)
                    except Exception as e:
                        print(f"Warning: Failed to add chart {idx}: {str(e)}")
            
            finally:
                # Clean up temporary files
                for chart_path in chart_paths:
                    try:
                        os.unlink(chart_path)
                    except Exception:
                        pass
        
        if include_stats:
            stats = self.calculate_summary_stats()
            ws_stats = wb.create_sheet("Summary Stats")
            
            # Add stats to worksheet
            for idx, (key, value) in enumerate(stats.items(), start=1):
                ws_stats[f'A{idx}'] = key
                ws_stats[f'B{idx}'] = value
                if isinstance(value, (int, float)):
                    value = format_currency(value)
                    ws_stats[f'B{idx}'].number_format = '"$"#,##0.00'
            
            apply_excel_styles(ws_stats)
        
        # Save workbook in binary mode
        try:
            wb.save(filename=output_file)
            wb.close()
        except Exception as e:
            raise Exception(f"Failed to save Excel file: {str(e)}")